import csv
import os

import numpy as np
from openpyxl import load_workbook

import constants
from model import Vc
from model import Vh


def read_input():
    combined_input = []
    for i in range(1, constants.NUMBER_OF_STUDENTS + 1):
        mode_input = []
        for j in range(1, constants.NUMBER_OF_MODES + 1):
            file_path = 'data/mode' + `j` + '/Student' + `i` + 'Mode' + `j` + '/'
            file_size = 0
            max_file_path = ''
            for file in os.listdir(file_path):
                if '.txt' in file:
                    curr_file_path = file_path + file
                    size = os.path.getsize(curr_file_path)
                    if file_size < size:
                        file_size = size
                        max_file_path = curr_file_path

            mode_input.append(readDataFile(max_file_path))
        combined_input.append(mode_input)
    return combined_input


def readDataFile(file_path=''):
    """
    Read file line be line and add to 2-dim arr
    """
    columns = []
    colNos = [4, 6, 10, 11, 12, 13, 25, 30, 31]
    with open(file_path) as fp:
        fp.readline()
        line = fp.readline()
        while line and len(line) > 0:
            items = []
            cols = line.split('\t')
            for colNo in colNos:
                if cols[colNo] is '-':
                    items.append(0)
                else:
                    items.append(float(cols[colNo]))
            vcobj = Vc(items[0], items[1], items[2], items[3], items[4], items[5], items[6], items[7], items[8])
            columns.append(vcobj)
            line = fp.readline()
    return columns


def split_data_sections(data, n):
    section = []

    last = 0
    i = 0
    rem = len(data) % n
    padding_size = n - rem
    section_size = len(data) / n
    if rem < n / 2:
        padding_size = 0
    else:
        section_size += 1

    add_padding(data, padding_size)

    while i < n:
        single_section = data[int(last):int(last + section_size)]
        section.append(get_average(single_section))
        last = last + section_size
        i += 1
    return section


def add_padding(data, padding_size):
    empty_vc = Vc(0, 0, 0, 0, 0, 0, 0, 0, 0)
    while padding_size != 0:
        data.append(empty_vc)
        padding_size = padding_size - 1


def get_average(section):
    vc_mean = Vc(0, 0, 0, 0, 0, 0, 0, 0, 0)
    for single_object in section:
        if single_object != 0:
            vc_mean = vc_mean + single_object

    vc_mean = vc_mean / len(section)

    return vc_mean


def getVC():
    total_vc = read_input()
    final_vc = []
    # split vc in groups of 6 for students 1-5
    for i in range(0, 1):
        student = total_vc[i]
        for j in range(0, constants.NUMBER_OF_MODES):
            final_vc.append(split_data_sections(student[j], 6))

    # split vc in groups of 5 for students 6, 7, 8, 10
    for i in range(5, 8):
        student = total_vc[i]
        for j in range(0, constants.NUMBER_OF_MODES):
            final_vc.append(split_data_sections(student[j], 5))

    final_vc.append(split_data_sections(total_vc[9][0], 5))
    final_vc.append(split_data_sections(total_vc[9][1], 5))
    final_vc.append(split_data_sections(total_vc[9][2], 5))

    # split vc in groups of 4 for student 9 mode 1 and 5 parts for student 9 mode 2 and mode 3.
    final_vc.append(split_data_sections(total_vc[8][0], 4))
    final_vc.append(split_data_sections(total_vc[8][1], 5))
    final_vc.append(split_data_sections(total_vc[8][2], 5))

    # split vc in groups of 6 for students 11-30
    for i in range(10, constants.NUMBER_OF_STUDENTS):
        student = total_vc[i]
        for j in range(0, constants.NUMBER_OF_MODES):
            final_vc.append(split_data_sections(student[j], 6))

    return final_vc


def getVH():
    """
    Read from output sheet (.xlsx)
    You need openpyxl package
    """
    filename = "data/Vh_Output_Data.xlsx"
    wb = load_workbook(filename)
    sheet = wb.active
    max_row = sheet.max_row
    max_column = sheet.max_column
    res = []
    user_num = 0
    for i in range(2, max_row + 1):
        if sheet.cell(row=i, column=max_column).value == user_num + 1:
            user_num = user_num + 1
        items = []
        for j in range(1, max_column):
            cell_obj = sheet.cell(row=i, column=j)
            items.append(cell_obj.value)

        res.append(Vh(items[0], items[1], items[2], items[3], items[4], user_num))
    return res


def mergeAndWriteToCsv(vc_list, vh_list):
    with open(constants.OUTPUT_FILE, 'wb') as my_file:
        wr = csv.writer(my_file, quoting=csv.QUOTE_MINIMAL)
        wr.writerow([' ', constants.VELOCITY, constants.LANE_POS, constants.SPEED, constants.STEER, constants.ACCEL,
                     constants.BRAKE, constants.LONG_ACCEL, constants.HEADWAY_TIME, constants.HEADWAY_DIST,
                     constants.USER, constants.MODE,
                     constants.SPEED, constants.NOE, constants.RESPONSE_TIME, constants.NOS])
        i = 0
        for vc, vh in zip(vc_list, vh_list):
            wr.writerow(
                [i, vc.velocity, vc.lanepos, vc.speed, vc.steer, vc.accel, vc.brake, vc.longAccel, vc.headwayTime,
                 vc.headwayDist,
                 vh.student, vh.mode, vh.speed, vh.noe, vh.responseTime, vh.nos])
            i += 1


def main():
    vc = getVC()
    vh_list = getVH()
    vc_list = []
    for student_mode in vc:
        for row in student_mode:
            vc_list.append(row)
    mergeAndWriteToCsv(vc_list, vh_list)


main()
